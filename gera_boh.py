################################################################################################
#			HEADER
#
# AUTOR:
#	Stenio Silveira <stenioas@gmail.com>
# NOME DA APLICAÇÃO:
#	gera_boh
# VERSÃO:
#	2.0
# DESCRIÇÃO:
#	Gera o Boletim de Ocupação Hoteleira do mês e ano escolhidos
#
################################################################################################

# IMPORTA AS BIBLIOTECAS EXTERNAS

from bs4 import BeautifulSoup
import mechanize
from win32com import client
import win32api
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# IMPORTA AS BIBLIOTECAS NATIVAS

import pathlib
from datetime import datetime
from datetime import timedelta
import getpass
import os
import platform
import calendar

# DADOS DA EMPRESA

company_url_name = "0"
company_name = "0"
num_embratur = "0"
num_of_rooms = "0"
num_of_beds = "0"

# PARÂMETROS INICIAIS

app_version = "1.0"
url_login = "https://pms.hospedin.com/login"
sistema = str(platform.system())
limpa_tela = ""
if sistema == "Windows":
	limpa_tela = "cls"
else:
	limpa_tela = "clear"

os.system(limpa_tela)

# FUNÇÕES

def bem_vindo():
	print("="*80)
	print("- BEM VINDO AO GERA-BOH - ".center(80))
	print("Script para Automatização do BOH".center(80))
	print("- HOSPEDIN -".center(80))
	print("="*80)
	print()

# ESCOLHENDO O MÊS E O ANO

bem_vindo()
mes_boh = input(str("Digite o mês desejado em formato MM: "))
mes_boh_int = int(mes_boh)
ano_boh = input(str("Digite o ano desejado em formato AAAA: "))
ano_boh_int = int(ano_boh)
dia_boh = '01'
os.system(limpa_tela)

# TELA DE LOGIN / SENHA

bem_vindo()
user = input(str("E-mail: "))
pwd = getpass.getpass("Senha: ")
os.system(limpa_tela)
print()
print(" Iniciando aplicação...")
print()

# DETECTA QUANTOS DIAS TEM O MÊS ESCOLHIDO

range_mes = calendar.monthrange(ano_boh_int,mes_boh_int)
dias_mes = range_mes[1]
dias_mes_str = str(dias_mes)

# INICIALIZA O BROWSER

br = mechanize.Browser()
br.set_handle_robots(False)
br.addheaders = [("User-agent","Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.2.13) Gecko/20101206 Ubuntu/10.10 (maverick) Firefox/3.6.13")]

# ABRINDO A PÁGINA DO HOSPEDIN

br.open(url_login)

# EFETUA LOGIN COM USUARIO E SENHA FORNECIDOS

br.select_form(nr=0)
br["user[email]"] = user
br["user[password]"] = pwd
br.submit()

# 	INICIALIZA A PORCENTAGEM DE PROGRESSO

pbar = int(0)
def pbarr_update(x, y):
	os.system(limpa_tela)
	print()
	print(x)
	print()
	print(" " + str(y) + "%")

# CARREGA O ARQUIVO XLSX NA MEMÓRIA PARA POSTERIOR INSERÇÃO DE DADOS

wb = load_workbook(filename = 'boh.xlsx')
ws = wb['RASCUNHO']

# LIMPA O ARQUIVO XLSX

ws['B1'] = 0
ws['B2'] = 0
ws['B3'] = 0
ws['C1'] = 0

for i in range(31):
	ws['B' + str(i + 6)] = 0
	ws['C' + str(i + 6)] = 0
	ws['D' + str(i + 6)] = 0

# COLETA A OCUPAÇÃO NO ÚLTIMO DIA DO MÊS ANTERIOR

pbar = pbar + 1
pbarr_update(' Coletando ocupação do último dia do mês anterior...', pbar)
dia_boh = "01"
data_ocupacao = str(dia_boh + "/" + mes_boh + "/" + ano_boh)
datai_increase = datetime.strptime(data_ocupacao, '%d/%m/%Y')
datai_increase = datai_increase - timedelta(days=1)
diai_validate = str(datai_increase.strftime('%d'))
mesi_validate = str(datai_increase.strftime('%m'))
anoi_validate = str(datai_increase.strftime('%Y'))
dataf_increase = datai_increase + timedelta(days=1095)
diaf_validate = str(dataf_increase.strftime('%d'))
mesf_validate = str(dataf_increase.strftime('%m'))
anof_validate = str(dataf_increase.strftime('%Y'))

# ACESSA A NOVA URL

url= br.open("https://pms.hospedin.com/" + company_url_name + "/reservations?search=&places%5Bplace_type_id%5D=&reservations%5Bsale_channel_id%5D=&date_field=check_out&date=" + dia_boh + "%2F" + mes_boh + "%2F" + ano_boh + "+-+" + diaf_validate + "%2F" + mesf_validate + "%2F" + anof_validate + "&between=" + diai_validate + "%2F" + mesi_validate + "%2F" + anoi_validate + "+-+" + diai_validate + "%2F" + mesi_validate + "%2F" + anoi_validate + "&reservations%5Bis_any_debit_open%5D=&reservations%5Bstatus%5D%5B%5D=3&reservations%5Bstatus%5D%5B%5D=4&reservations%5Bstatus%5D%5B%5D=5")

# PREPARA OS DADOS PARA COLETA

soup = BeautifulSoup(url, 'html5lib')

# TESTA O VALOR DOS ADULTOS, SE NÃO EXISTIR RETORNA '0', CASO EXISTA COLETA O VALOR

try:
	elmnt = soup.find_all('span', attrs={"class": "blue-grey lighten-3 text-bold-300 ml-1"})
	elmnt_txt = str(elmnt[0])
	adultos_uma = int(elmnt_txt.split(" ")[34])
except Exception:
	adultos_uma = 0

# TESTA O VALOR DAS CRIANÇAS, CASO NÃO EXISTA RETORNA '0', CASO EXISTA ATRIBUI O VALOR

try:
	elmnt = soup.find_all('span', attrs={"class": "blue-grey lighten-3 text-bold-300 ml-1"})
	elmnt_txt = str(elmnt[0])
	criancas_uma = int(elmnt_txt.split(" ")[50])
except Exception:
	criancas_uma = 0

total_hospedes_uma = adultos_uma + criancas_uma

# ENVIA O VALOR PARA A PLANILHA

ws['B3'] = total_hospedes_uma

# COLETA AS ENTRADAS EM CADA DIA DO MÊS

contador_entradas = int(1)
cont_linha_entrada = int(6)
while(contador_entradas <= dias_mes):

	dia_boh = ""
	if contador_entradas <= 9:
		dia_boh = str(contador_entradas)
		dia_boh = ("0" + dia_boh)
	else:
		dia_boh = str(contador_entradas)

	url = br.open("https://pms.hospedin.com/" + company_url_name + "/reservations?search=&places%5Bplace_type_id%5D=&reservations%5Bsale_channel_id%5D=&date_field=check_in&date=" + dia_boh + "%2F" + mes_boh + "%2F" + ano_boh + "+-+" + dia_boh + "%2F" + mes_boh + "%2F" + ano_boh + "&between=&reservations%5Bis_any_debit_open%5D=&reservations%5Bstatus%5D%5B%5D=3&reservations%5Bstatus%5D%5B%5D=4&reservations%5Bstatus%5D%5B%5D=5")

	# PREPARA OS DADOS PARA COLETA

	soup = BeautifulSoup(url, 'html5lib')

	# TESTA O VALOR DOS ADULTOS, CASO NÃO EXISTA RETORNA '0', CASO EXISTA ATRIBUI O VALOR

	try:
		elmnt = soup.find_all('span', attrs={"class": "blue-grey lighten-3 text-bold-300 ml-1"})
		elmnt_txt = str(elmnt[0])
		adultos = int(elmnt_txt.split(" ")[34])
	except Exception:
		adultos = 0

	# TESTA O VALOR DAS CRIANÇAS, CASO NÃO EXISTA RETORNA '0', CASO EXISTA ATRIBUI O VALOR

	try:
		elmnt = soup.find_all('span', attrs={"class": "blue-grey lighten-3 text-bold-300 ml-1"})
		elmnt_txt = str(elmnt[0])
		criancas = int(elmnt_txt.split(" ")[50])
	except Exception:
		criancas = 0

	total_hospedes = adultos + criancas

	# ENVIA O TOTAL DE HÓSPEDES QUE ENTRARAM NO DIA PARA A PLANILHA

	linha_entrada = str(cont_linha_entrada)
	ws['B' + linha_entrada] = total_hospedes
	
	contador_entradas = contador_entradas + 1
	cont_linha_entrada = cont_linha_entrada + 1
	pbar = pbar + 1
	pbarr_update(' Coletando as entradas...', pbar)

# COLETA AS SAÍDAS EM CADA DIA DO MÊS

contador_saidas = int(1)
cont_linha_saida = int(6)
while(contador_saidas <= dias_mes):

	dia_boh = ""
	if contador_saidas <= 9:
		dia_boh = str(contador_saidas)
		dia_boh = ("0" + dia_boh)
	else:
		dia_boh = str(contador_saidas)

	url = br.open("https://pms.hospedin.com/" + company_url_name + "/reservations?search=&places%5Bplace_type_id%5D=&reservations%5Bsale_channel_id%5D=&date_field=check_out&date=" + dia_boh + "%2F" + mes_boh + "%2F" + ano_boh + "+-+" + dia_boh + "%2F" + mes_boh + "%2F" + ano_boh + "&between=&reservations%5Bis_any_debit_open%5D=&reservations%5Bstatus%5D%5B%5D=3&reservations%5Bstatus%5D%5B%5D=4&reservations%5Bstatus%5D%5B%5D=5")

	# PREPARA OS DADOS PARA COLETA

	soup = BeautifulSoup(url, 'html5lib')

	# TESTA O VALOR DOS ADULTOS, CASO NÃO EXISTA RETORNA '0', CASO EXISTA ATRIBUI O VALOR

	try:
		elmnt = soup.find_all('span', attrs={"class": "blue-grey lighten-3 text-bold-300 ml-1"})
		elmnt_txt = str(elmnt[0])
		adultos = int(elmnt_txt.split(" ")[34])
	except Exception:
		adultos = 0

	# TESTA O VALOR DAS CRIANÇAS, CASO NÃO EXISTA RETORNA '0', CASO EXISTA ATRIBUI O VALOR

	try:
		elmnt = soup.find_all('span', attrs={"class": "blue-grey lighten-3 text-bold-300 ml-1"})
		elmnt_txt = str(elmnt[0])
		criancas = int(elmnt_txt.split(" ")[50])
	except Exception:
		criancas = 0

	total_hospedes = adultos + criancas

	# ENVIA O TOTAL DE HÓSPEDES QUE SAIRAM NO DIA PARA A PLANILHA
	
	linha_saida = str(cont_linha_saida)
	ws['C' + linha_saida] = total_hospedes

	contador_saidas = contador_saidas + 1
	cont_linha_saida = cont_linha_saida + 1
	pbar = pbar + 1
	pbarr_update(' Coletando as saídas...', pbar)

# COLETA A OCUPAÇÃO EM CADA DIA DO MÊS

contador_ocupacao = 1
cont_linha_ocupacao = 6
while(contador_ocupacao <= dias_mes):

	dia_boh = ""
	if contador_ocupacao <= 9:
		dia_boh = str(contador_ocupacao)
		dia_boh = ("0" + dia_boh)
	else:
		dia_boh = str(contador_ocupacao)

	url = br.open("https://pms.hospedin.com/" + company_url_name + "/reports/occupancy_reports?guests%5Bcompany_id%5D=&reservations%5Bguest_id%5D=&reservations%5Bplace_type_id%5D=&reservations%5Bplace_id%5D=&period=custom&date_action=&perspective=by_period&begin_date=" + dia_boh + "%2F" + mes_boh + "%2F" + ano_boh + "&end_date=" + dia_boh + "%2F" + mes_boh + "%2F" + ano_boh + "&old_perspective=by_period&must_keep_date=false")
	
	# PREPARA OS DADOS PARA COLETA

	soup = BeautifulSoup(url, 'html5lib')

	# TESTA O VALOR DA OCUPAÇÃO, SE NÃO EXISTIR RETORNA '0'

	try:
		elmnt = soup.find_all('td')
		elmnt_txt = str(elmnt)
		elmnt_ocup = elmnt_txt.split(">")[7]
		if len(elmnt_ocup) == 5:
			ocupacao = int(elmnt_ocup[0:1])
		else:
			ocupacao = int(elmnt_ocup[0:2])
	except Exception:
		ocupacao = 0

	# ENVIA A OCUPAÇÃO PARA A PLANILHA
	
	linha_ocupacao = str(cont_linha_ocupacao)
	if ocupacao <= 23:
		ws['D' + linha_ocupacao] = ocupacao
	else:
		ws['D' + linha_ocupacao] = 23

	contador_ocupacao = contador_ocupacao + 1
	cont_linha_ocupacao = cont_linha_ocupacao + 1
	pbar = pbar + 1
	pbarr_update(' Coletando a ocupação...', pbar)

# INSERE OS ULTIMOS DADOS

ws['B1'] = int(mes_boh)
ws['C1'] = int(ano_boh)
ws['B2'] = int(dias_mes_str)
ws['F1'] = company_name
ws['F2'] = company_url_name
ws['F3'] = num_embratur
ws['F4'] = num_of_rooms
ws['F5'] = num_of_beds

# SALVA E FECHA A PLANILHA COM MÊS E ANO

try:
	wb.save("boh.xlsx")
	wb.close()
except Exception as e:
	print(" Ocorreu um erro ao salvar a pasta de trabalho do BOH, tente novamente ou reinicie o computador!")

# MÊS POR EXTENSO

if mes_boh == "01":
	mes_boh_extenso = "JAN"
elif mes_boh == "02":
	mes_boh_extenso = "FEV"
elif mes_boh == "03":
	mes_boh_extenso = "MAR"
elif mes_boh == "04":
	mes_boh_extenso = "ABR"
elif mes_boh == "05":
	mes_boh_extenso = "MAI"
elif mes_boh == "06":
	mes_boh_extenso = "JUN"
elif mes_boh == "07":
	mes_boh_extenso = "JUL"
elif mes_boh == "08":
	mes_boh_extenso = "AGO"
elif mes_boh == "09":
	mes_boh_extenso = "SET"
elif mes_boh == "10":
	mes_boh_extenso = "OUT"
elif mes_boh == "11":
	mes_boh_extenso = "NOV"
elif mes_boh == "12":
	mes_boh_extenso = "DEZ"

# GERANDO O ARQUIVO PDF

pbar = pbar + 1
pbarr_update('Gerando PDF...', pbar)

data_exec_orig = datetime.now()
data_exec = data_exec_orig.strftime('%d.%m.%Y-%H.%M.%S')
excel_file = "boh.xlsx"
pdf_file = "BOH - " + company_name + " - " + mes_boh_extenso + " " + ano_boh + " - ger." + data_exec + ".pdf"
excel_path = str(pathlib.Path.cwd() / excel_file)
pdf_path = str(pathlib.Path.cwd() / 'Arquivo' / pdf_file)

# --- CRIA O DIRETORIO ARQUIVO CASO NÃO EXISTA
if os.path.exists(str(pathlib.Path.cwd() / 'Arquivo')):
	pass
else:
	os.system("md Arquivo")

excel = client.DispatchEx("Excel.Application")
excel.Visible = 0
wb_pdf = excel.Workbooks.Open(excel_path)
ws_pdf = wb_pdf.Worksheets['BOH ATUAL']
try:
	ws_pdf.SaveAs(pdf_path, FileFormat=57)
except Exception as e:
	print(" Erro ao gerar o PDF...")
	print(str(e))
finally:
	wb_pdf.Save()
	wb_pdf.Close()
	excel.Quit()

# FINALIZANDO

if pbar < 100:
	for i in range(100 - pbar):
		pbar = pbar + 1
		pbarr_update(' Finalizando...', pbar)

print()
print(" Sucesso!")
